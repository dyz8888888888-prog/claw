#!/usr/bin/env python3
"""Export convertible bond concept list to Excel"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# Load data
with open('C:/Users/DYZ/WorkBuddy/Claw/可转债日报/cb_concept_map.json', encoding='utf-8') as f:
    cb_map = json.load(f)
with open('C:/Users/DYZ/WorkBuddy/Claw/可转债日报/cb_concept_heat.json', encoding='utf-8') as f:
    heat = json.load(f)
with open('C:/Users/DYZ/WorkBuddy/Claw/cb_monitor/data/concept_alias.json', encoding='utf-8') as f:
    alias = json.load(f)

DATA_DIR = 'C:/Users/DYZ/WorkBuddy/Claw/cb_monitor/outputs'
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT = os.path.join(DATA_DIR, f'可转债概念列表_{date.today().isoformat()}.xlsx')

# Build reverse map: concept -> CB list
concept_to_cb = {}
for cb_code, info in cb_map.items():
    for c in info.get('concepts', []):
        concept_to_cb.setdefault(c, []).append({
            'code': cb_code,
            'name': info.get('name', ''),
        })

wb = Workbook()

# --- Styles ---
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
data_font = Font(name='Microsoft YaHei', size=10)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
alt_fill = PatternFill(start_color='F5F8FC', end_color='F5F8FC', fill_type='solid')

def style_header(ws, cols, row=1):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_data(ws, max_row, max_col, start_row=2):
    for r in range(start_row, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill

# ===== Sheet 1: CB -> Concepts =====
ws1 = wb.active
ws1.title = '转债-概念映射'
ws1.append(['转债代码', '转债名称', '概念数', '概念列表'])
for cb_code, info in sorted(cb_map.items()):
    concepts = info.get('concepts', [])
    ws1.append([cb_code, info.get('name', ''), len(concepts), ', '.join(concepts)])

style_header(ws1, 4)
style_data(ws1, ws1.max_row, 4)
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 80
ws1.auto_filter.ref = f'A1:D{ws1.max_row}'
ws1.freeze_panes = 'A2'

# ===== Sheet 2: Concept -> CBs =====
ws2 = wb.create_sheet('概念-转债')
ws2.append(['概念名称', '转债数', '热度分', '转债列表(代码+名称)'])
for concept, cbs in sorted(concept_to_cb.items(), key=lambda x: -len(x[1])):
    h = heat.get(concept, 0)
    cb_list = ', '.join([f"{cb['code']} {cb['name']}" for cb in cbs])
    ws2.append([concept, len(cbs), h, cb_list])

style_header(ws2, 4)
style_data(ws2, ws2.max_row, 4)
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 100
ws2.auto_filter.ref = f'A1:D{ws2.max_row}'
ws2.freeze_panes = 'A2'

# Highlight heat scores
for r in range(2, ws2.max_row+1):
    h_val = ws2.cell(row=r, column=3).value or 0
    if h_val >= 30:
        ws2.cell(row=r, column=3).font = Font(name='Microsoft YaHei', size=10, color='EF4444', bold=True)
    elif h_val >= 20:
        ws2.cell(row=r, column=3).font = Font(name='Microsoft YaHei', size=10, color='F59E0B', bold=True)

# ===== Sheet 3: Concept Heat Ranking =====
ws3 = wb.create_sheet('概念热度排行')
ws3.append(['排名', '概念名称', '热度分', '转债数', '热度等级'])
for rank, (concept, h) in enumerate(sorted(heat.items(), key=lambda x: -x[1]), 1):
    cnt = len(concept_to_cb.get(concept, []))
    level = '高' if h >= 30 else ('中' if h >= 15 else ('低' if h >= 5 else '冷'))
    ws3.append([rank, concept, h, cnt, level])

style_header(ws3, 5)
style_data(ws3, ws3.max_row, 5)
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 10
ws3.auto_filter.ref = f'A1:E{ws3.max_row}'
ws3.freeze_panes = 'A2'

for r in range(2, ws3.max_row+1):
    level = ws3.cell(row=r, column=5).value
    if level == '高':
        ws3.cell(row=r, column=5).font = Font(name='Microsoft YaHei', size=10, color='EF4444', bold=True)
        ws3.cell(row=r, column=3).font = Font(name='Microsoft YaHei', size=10, color='EF4444', bold=True)
    elif level == '中':
        ws3.cell(row=r, column=5).font = Font(name='Microsoft YaHei', size=10, color='F59E0B', bold=True)

# ===== Sheet 4: Concept Alias Map =====
ws4 = wb.create_sheet('概念别名映射')
ws4.append(['Fuyao原始概念', '标准概念名', '备注'])
alias_map = alias.get('alias', {})
for fuyao_name, std_name in sorted(alias_map.items()):
    heat_val = heat.get(std_name, 0)
    cb_cnt = len(concept_to_cb.get(std_name, []))
    note = f'热度{heat_val}, {cb_cnt}只转债'
    ws4.append([fuyao_name, std_name, note])

style_header(ws4, 3)
style_data(ws4, ws4.max_row, 3)
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 25

# ===== Sheet 5: Summary =====
ws5 = wb.create_sheet('汇总')
summary_rows = [
    ['项目', '数值'],
    ['转债总数', len(cb_map)],
    ['独特概念数', len(concept_to_cb)],
    ['有热度评分的概念', len(heat)],
    ['概念别名数', len(alias_map)],
    ['数据来源', 'i问财缓存 + TDX补充'],
    ['导出日期', date.today().isoformat()],
    ['', ''],
    ['热度等级说明', ''],
    ['高 (>=30)', '芯片概念(39.5) 新能源汽车(45.4) 储能(40.5) 融资融券(100) 等'],
    ['中 (15-30)', '光伏概念(28.4) 锂电池(25.0) 军工(19.3) 等'],
    ['低 (5-15)', '冷链物流(8.2) 工业大麻(5.5) 等'],
    ['冷 (<5)', '氧化铝(0.2) 等'],
    ['', ''],
    ['Top 10 概念(按转债数)', ''],
]
for c, cnt in sorted(concept_to_cb.items(), key=lambda x: -len(x[1]))[:10]:
    h = heat.get(c, 0)
    summary_rows.append([f'{c} ({cnt}只)', f'热度 {h}'])

summary_rows.append(['', ''])
summary_rows.append(['Top 10 概念(按热度)', ''])
for c, h in sorted(heat.items(), key=lambda x: -x[1])[:10]:
    cnt = len(concept_to_cb.get(c, []))
    summary_rows.append([f'{c} (热度{h})', f'{cnt}只转债'])

for row in summary_rows:
    ws5.append(row)

ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 35
for r in range(1, ws5.max_row+1):
    ws5.cell(row=r, column=1).font = data_font
    ws5.cell(row=r, column=2).font = data_font

wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'Sheets: {wb.sheetnames}')
print(f'Size: {os.path.getsize(OUTPUT)/1024:.1f} KB')
